"""
Table extraction service — LLM-powered structured data extraction from documents.

Flow:
1. Retrieve ALL chunks for a document (uses Phase 2's get_all_chunks_for_document)
2. Prompt LLM to extract the described table as JSON
3. Validate the LLM response is valid JSON (with one retry if not)
4. Convert JSON rows → .xlsx bytes via openpyxl

This is the killer feature for CMMS/PMS data entry: take a multilingual
maritime parts catalog PDF, get a spreadsheet of parts ready for import.
"""

import io
import json
import logging
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.services.retrieval.llm import LLMService
from app.services.retrieval.vectordb import VectorDBService

logger = logging.getLogger(__name__)


EXTRACTION_PROMPT = """You are extracting structured data from a maritime technical document.

The user wants to extract: {description}

Below is the COMPLETE content of one document. Read it carefully and produce a JSON array
of objects representing the table the user described.

Rules:
1. Output ONLY valid JSON — no commentary, no markdown fences, no explanation.
2. Start your response with `[` and end with `]`. Nothing else.
3. Each object in the array is one row. Use consistent keys across all objects.
4. Include EVERY row present in the document. If the document has 57 parts, output 57 objects.
5. Use snake_case keys derived from column headers (e.g. "part_no", "designation", "quantity").
6. Quote numeric values exactly as they appear (don't auto-convert types).
7. If a cell is empty, use null (not "").
8. For multilingual labels (e.g. German/English/French part names), include all in one
   "designation" field separated by " / " unless the user asked otherwise.

<document>
{context}
</document>

Output the JSON array now:"""


class ExtractionService:
    """Extract structured tables from documents and return as Excel files."""

    def __init__(self):
        self.vectordb = VectorDBService()
        self.llm = LLMService()

    def extract_to_excel(
        self,
        document_id: str,
        tenant_id: str,
        description: str,
    ) -> tuple[bytes, str]:
        """
        Extract a table from a document and return (xlsx_bytes, suggested_filename).

        Raises ValueError if the document has no content or the LLM returns
        unparseable JSON after one retry.
        """
        chunks = self.vectordb.get_all_chunks_for_document(
            document_id=document_id, tenant_id=tenant_id
        )
        if not chunks:
            raise ValueError("Document has no indexed content yet.")

        doc_title = chunks[0].get("title", "document")
        context = self._build_context(chunks)

        rows = self._call_llm_and_parse(description, context)

        if not rows:
            raise ValueError("The LLM did not find any matching rows to extract.")

        xlsx_bytes = self._build_xlsx(rows, sheet_name=self._safe_sheet_name(description))
        filename = f"{self._safe_filename(doc_title)}_{self._safe_filename(description)}.xlsx"
        return xlsx_bytes, filename

    def _build_context(self, chunks: list[dict]) -> str:
        sections = []
        for chunk in chunks:
            page = chunk.get("page_number", "?")
            sections.append(f"[Page {page}]\n{chunk['text']}")
        return "\n\n---\n\n".join(sections)

    def _call_llm_and_parse(self, description: str, context: str) -> list[dict]:
        """Call LLM, parse JSON. Retries once with a clarifying nudge if invalid."""
        prompt = EXTRACTION_PROMPT.format(description=description, context=context)

        # First attempt
        response = self.llm.get_answer(question=prompt, context="")
        parsed = self._try_parse_json(response)
        if parsed is not None:
            return parsed

        # Retry with explicit nudge
        logger.warning("First extraction attempt produced invalid JSON; retrying")
        retry_prompt = (
            prompt
            + "\n\nIMPORTANT: Your previous response was not valid JSON. "
            "Reply with ONLY the JSON array, starting with `[` and ending with `]`. "
            "Do not include any other text, code fences, or explanation."
        )
        response = self.llm.get_answer(question=retry_prompt, context="")
        parsed = self._try_parse_json(response)
        if parsed is None:
            raise ValueError(
                "Could not extract structured data. The LLM did not return valid JSON. "
                "Try rephrasing what you want extracted, e.g. 'parts list with part number, "
                "designation, and quantity'."
            )
        return parsed

    def _try_parse_json(self, raw: str) -> list[dict] | None:
        """Tolerate code-fenced JSON and stray text around the array."""
        if not raw:
            return None
        # Strip markdown code fences if present
        fenced = re.search(r"```(?:json)?\s*(\[[\s\S]*?\])\s*```", raw)
        if fenced:
            raw = fenced.group(1)
        else:
            # Grab the first [...] block
            match = re.search(r"\[[\s\S]*\]", raw)
            if match:
                raw = match.group(0)
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            return None
        if not isinstance(data, list):
            return None
        # Coerce non-dict items into {value: ...} so we always have rows
        return [item if isinstance(item, dict) else {"value": item} for item in data]

    def _build_xlsx(self, rows: list[dict], sheet_name: str) -> bytes:
        """Convert list-of-dicts → .xlsx bytes with styled header row."""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Collect column order: keys from first row + any new keys discovered later
        columns: list[str] = []
        seen: set[str] = set()
        for row in rows:
            for k in row.keys():
                if k not in seen:
                    seen.add(k)
                    columns.append(k)

        # Header row
        header_fill = PatternFill("solid", fgColor="0F2B5B")  # NAUTOS navy
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row.get(col_name)
                # openpyxl can't write dicts/lists directly — stringify
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Reasonable column widths (cap at 50 chars to avoid massive columns)
        for col_idx, col_name in enumerate(columns, start=1):
            max_len = len(str(col_name))
            for row in rows:
                v = row.get(col_name)
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

        # Freeze header row
        ws.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        """Excel sheet names: max 31 chars, no `/ \\ * ? : [ ]`."""
        cleaned = re.sub(r"[/\\*?:\[\]]", "", name).strip() or "Sheet1"
        return cleaned[:31]

    @staticmethod
    def _safe_filename(name: str) -> str:
        return re.sub(r"[^a-zA-Z0-9_-]+", "_", name).strip("_") or "extract"
